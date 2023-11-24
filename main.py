"""
Taha Ghumman, Troy Williams, Rashed Amani
LEC 1D02
Nov 24th, 2023
--

This program aims to cleanse the data necessary to report the amount of collisions based off different categories.
The program focuses on finding which animal collided the most, which year the most incidents happened, which month the most incidents happend, and which airline/operator caused the most incidents.
This is done by finding the needed columns of each focus, then counting the occurences of the value of each cell in the column.
The program takes in different aircraft data related to incident reports as an excel/.xlsx file to read through it. 
The program relies on the OpenPyXL library to help read the necessary data.

References:
Python Openpyxl, JavaTpoint, 2021. Python Openpyxl Tutorial - javatpoint
Hands-on Python Openpyxl Tutorial With Examples (softwaretestinghelp.com)

Python Excel, 2021. Openpyxl Tutorial - Read, Write & Manipulate xlsx files in Python -
Python Excel

Contributions:

The data cleansing, finding a species common name, counting the data and main function were created by Taha Ghumman.

The creation of the excel file, writing to the excel file, alphabetical sorting and graph creation were done by Troy Williams.

The value deduction of the top values and maximum values of each column were made by Rashed Amani.

"""

import openpyxl as pyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
import re

workBookPath = "mediumAircraftData.xlsx"
newDataSheetPath = r"wildlifeAnalysis.xlsx"
print("Loading workbook...")

workBook = pyxl.load_workbook(workBookPath)
print("Workbook loaded, cleaning data...")

workSheet = workBook.active

wildlifeAnalysis = Workbook()

# Getting Species Name - Taha Ghumman
def getSpeciesName(animalName: str):
    """
    Given the "animalName" specified, it will strip the full name of the animal into its common name.
    If there is multiple animals it will make sure to separate each animal and account for them as individual animals.
    The function will also make sure that if the animal name ends with an "s" or is a plural of an animal - such as geese - to fix the plurality.
    The function does this by using regex to allow it to remove all unneccessary items that do not need to be parsed.

    :params animalName: The name of the animal being stripped of its full name down to the common name.
    :return splitAnimalName/animals: returns only the common name that it has found by stripping away the common name.
    """
    splitAnimalName = animalName.split(",")

    if len(splitAnimalName) == 1:
        splitAnimalName = re.split(r"[-;,.\s]\s*", splitAnimalName[0])
        
        if splitAnimalName[-1][-1] == "S":
            return splitAnimalName[-1][:-1]
        elif splitAnimalName[-1] == "GEESE":
            return "GOOSE"
        
        return splitAnimalName[-1]
    else:
        animals = []
        for animalNames in splitAnimalName:
            animalNames = re.split(r"[-;,.\s]\s*", animalNames)

            if animalNames[-1][-1] == "S":
                 animals.append(animalNames[-1][:-1])
            elif animalNames[-1] == "GEESE":
                animals.append("GOOSE")
            else:
                animals.append(animalNames[-1])
        return animals

# Counting All Animals - Taha Ghumman
def countAnimals(animalsWorksheet: dict):
    """
    Given the animal column, it will run the "getSpeciesName" function to strip the animal of its full name down to the common name.
    The code also makes sure to account for the blanks by filling them with UNKNOWN.
    The code will then loop through each cell in the column and tally up all the animals. If it encounters a new animal, it will make a new key in the animalDict.

    :params animalsWorksheet: The given column containing all the cells with animals.
    :return animalDict: returns the new counted animal column with how many incidents there are per animal. 
    """
    animalDict = {}
    for animalData in range(len(animalsWorksheet)): 
        fullName = animalsWorksheet[animalData].value

        if fullName == None or fullName == "": fullName = "UNKNOWN"
        
        commonName = getSpeciesName(fullName)

        if type(commonName) == list:
            for name in commonName:    
                if name not in animalDict:
                    animalDict[name] = 1
                else :
                    animalDict[name] += 1
        else:
            if commonName not in animalDict:
                animalDict[commonName] = 1
            else :
                animalDict[commonName] += 1

    return animalDict

# Counting All Items - Taha Ghumman
def countItems(someWorksheet: dict):
    """
    Counts the items given in a specified column. It will make sure that if a new data point is given, it will make it a new key.
    If the data point already exists, it will add onto the current value by one.

    :return countedDict: Returns a dictionary with the counted number for each key occurence in someWorksheet.
    """
    countedDict = {}
    for worksheetData in range(len(someWorksheet)):
        someValue = someWorksheet[worksheetData].value
        if someValue in countedDict:
            countedDict[someValue] += 1
        else:
            countedDict[someValue] = 1
    return countedDict

# Maximum Values and Top Values - Rashed Amani
def returnHighestandTop(someDict: dict, threshold: int = 0):
    """
    Returns the Highest values within of "someDict" as well as the top values within a given percentage threshold.
    The function accounts for ALL highest values even if there are many with the same maximum value. It does so by finding the key with the max value and then comparing with other keys and their values.
    The function finds the top values within a given percentage by seeing if they pass a threshold percentage. If they do, they are all considered to be within the percentage required to be appended to the list.

    :params someDict: The dictionary used to find its highest data points and the top data within a threshold.
            threshold: The threshold as a percentage used to see if values within the list are higher than that amount based off the highest given value. If no threshold is specified, it will assume you want all the values in the dictionary.
    :return highestDataPoints: The keys of the maximum values within of the given dictionary.
            topData: A list of tuples with its first value representing the key and the second value representing the value. The indices in the list all must be greater than the threshold percentage based on the maximum value.
    """
    highestData = max(someDict, key=someDict.get)
    highestDataPoints = []
    topData = []
    percentage = threshold / 100

    for dataPoint in someDict:
        if someDict[dataPoint] == someDict[highestData]:
            highestDataPoints.append(dataPoint)
            topData.append((dataPoint, someDict[dataPoint]))
            
        elif someDict[dataPoint] > (someDict[highestData] * percentage):
            topData.append((dataPoint, someDict[dataPoint]))
    
    return highestDataPoints, topData

# Creating Excel Sheet and Graphing - Troy Williams
def writeToExcel(dataList: list, name: str, itemName: str, chartTitle: str) :
    """
    Function that takes in a multi-dimensional list of data and sorts it alphabetically or ascending, (dependent on integer or string values.)
    The sorted list is then iterated through and placed in to a new excel sheet.
    Once the data placement is finished, the table of values is used to make a bar graph.
    :params dataList: multi-dimensional array with tuples consisting of a unique item and their occurrences within a dataset.
            name: name used for the new sheet being created for the data.
            itemName: name used for the unique items, header for column one, and title for the x-axis of the chart.
            chartTitle: specific name used for the title of the chart.
    """
    wildlifeAnalysis.create_sheet(name)
    sheet = wildlifeAnalysis[name]
    sortedList = sorted(dataList)
    
    sheet.cell(1,1,itemName)
    sheet.cell(1,2,"INCIDENTS")
    
    row,column = 2,1
    for item in sortedList :
        sheet.cell(row,column,item[0])
        sheet.cell(row,column+1,item[1])
        row += 1
        
    quantity = Reference(sheet, min_col=2, max_col=2, min_row=1, max_row=(len(sortedList)+1))
    items = Reference(sheet, min_col=1, max_col=1, min_row=2, max_row=(len(sortedList)+1))
    
    chart = BarChart()
    chart.add_data(quantity, titles_from_data = True)
    chart.set_categories(items)
    chart.title = chartTitle
    chart.x_axis.title = itemName
    chart.y_axis.title = "COLLISIONS"
    sheet.add_chart(chart,"D1")

# Main Function // Cleaning the Data - Taha Ghumman
def cleanData():
    """
    Main function body in which we grab each usable column from our main worksheet and name them as "xWorksheet", where x represents the data of the column.
    The function then grabs a dictionary where all the values are counted using the "countItems" and "countAnimals" function, for the number of occurences of their keys.
    The function is then grabbing the maximum values within the specified dictionaries for each column using "returnHighestandTop" as well as grabbing the top values in a given threshold. If there is no threshold specified; it will return all the values.
    The function then checks to see if there is more than 15 airlines, if there is, it will only get the ones within the 10% threshold of the highest value in the airlines dictionary.
    The function then prints each individual statistic with their highest value as well as the total incidents within that statistic.
    The function will then write each value it has grabbed from the calculated top values into an excel sheet and graph them into a bar graph based using the function "WriteToExcel".
    The function fixes the excel sheet by making a correction by removing the extra sheet the code starts with and saves all the sheet data into a directed data sheet path.
    """
    animalsWorksheet, yearWorksheet, monthWorksheet, airlineWorksheet = workSheet["AF"][1:], workSheet["B"][1:], workSheet["C"][1:], workSheet["F"][1:]

    animalDict, yearDict, monthDict, airlineDict = countAnimals(animalsWorksheet), countItems(yearWorksheet), countItems(monthWorksheet), countItems(airlineWorksheet)

    highestAnimals, topTenPercentAnimals = returnHighestandTop(animalDict, 10)
    highestYear, allYears = returnHighestandTop(yearDict)
    highestMonth, allMonths = returnHighestandTop(monthDict)
    highestAirline, topAirlines = returnHighestandTop(airlineDict)

    if len(airlineDict) > 15:
        highestAirline, topAirlines = returnHighestandTop(airlineDict, 10)

    print(f"The animal(s) mostly involved in incidents are: {highestAnimals} with {animalDict[highestAnimals[0]]} incidents total.")
    print(f"The year(s) with most accidents are: {highestYear}, with {yearDict[highestYear[0]]} incidents total")
    print(f"The month(s) with most accidents are: {highestMonth}, with {monthDict[highestMonth[0]]} incidents total")
    print(f"The airline(s) mostly involved in accidents are: {highestAirline}, with {airlineDict[highestAirline[0]]} incidents total")


    writeToExcel(topTenPercentAnimals, "ChartForAnimals", "ANIMALS", "Animals Involved in Aircraft Collisions")
    writeToExcel(topAirlines, "ChartForAirlines", "AIRLINES", "Airlines Involved in Aircraft Collisions")
    writeToExcel(allYears, "ChartForYears", "YEARS", "Aircraft Collisions by Year")
    writeToExcel(allMonths, "ChartForMonths", "MONTHS", "Aircraft Collisions by Month")

    del wildlifeAnalysis["Sheet"]
    wildlifeAnalysis.save(newDataSheetPath)
    
cleanData()
